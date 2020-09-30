from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier as KNN
from sklearn.metrics import matthews_corrcoef, confusion_matrix
from sklearn.metrics import classification_report as class_re
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from collections import namedtuple
import numpy as np
from os import path

def split_transform(X, Y, states=20):
    """Given X and Y returns a split and scaled version of them"""
    scaling = MinMaxScaler()
    esterase = ['EH51(22)', 'EH75(16)', 'EH46(23)', 'EH98(11)', 'EH49(23)']

    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=states, stratify=Y)

    X_train = X_train.loc[[x for x in X_train.index if x not in esterase]]
    X_test = X_test.loc[[x for x in X_test.index if x not in esterase]]
    Y_train = Y_train.loc[[x for x in Y_train.index if x not in esterase]]
    Y_test = Y_test.loc[[x for x in Y_test.index if x not in esterase]]

    transformed_x = scaling.fit_transform(X_train)
    transformed_x = pd.DataFrame(transformed_x)
    transformed_x.index = X_train.index
    transformed_x.columns = X_train.columns

    test_x = scaling.transform(X_test)
    test_x = pd.DataFrame(test_x)
    test_x.index = X_test.index
    test_x.columns = X_test.columns

    return transformed_x, test_x, Y_train, Y_test, X_train, X_test

def vote(pred1, pred2, pred3=None, pred4=None, pred5=None):
    """Hard voting for the ensembles"""
    vote_ = []
    index = []
    if pred3 is None:
        mean = np.mean([pred1, pred2], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            else:
                vote_.append(pred2[s])
                index.append(s)
    elif pred4 is None and pred5 is None:
        mean = np.mean([pred1, pred2, pred3], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            else:
                vote_.append(0)
                index.append(s)

    elif pred5 is None:
        mean = np.mean([pred1, pred2, pred3, pred4], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            elif x < 0.5:
                vote_.append(0)
                index.append(s)
            else:
                vote_.append(pred4[s])
                index.append(s)
    else:
        mean = np.mean([pred1, pred2, pred3, pred4], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            else:
                vote_.append(0)
                index.append(s)

    return vote_, index

def print_score(Y_test, y_grid, train_predicted, Y_train, test_index=None, train_index=None, mode=None):
    """ The function prints the scores of the models and the prediction performance """
    score_tuple = namedtuple("scores", ["test_confusion", "tr_report", "te_report",
                                        "train_mat", "test_mat", "train_confusion"])

    target_names = ["class 0", "class 1"]

    # looking at the scores of those predicted by al 3 of them
    if mode:
        Y_test = Y_test.iloc[[x for x in range(len(Y_test)) if x not in test_index]]
        Y_train = Y_train.iloc[[x for x in range(len(Y_train)) if x not in train_index]]
        y_grid = [y_grid[x] for x in range(len(y_grid)) if x not in test_index]
        train_predicted = [train_predicted[x] for x in range(len(train_predicted)) if x not in train_index]

    # Training scores
    train_confusion = confusion_matrix(Y_train, train_predicted)
    train_matthews = matthews_corrcoef(Y_train, train_predicted)
    # print(f"Y_train : {Y_train}, predicted: {train_predicted}")
    tr_report = class_re(Y_train, train_predicted, target_names=target_names, output_dict=True)

    # Test metrics
    test_confusion = confusion_matrix(Y_test, y_grid)
    test_matthews = matthews_corrcoef(Y_test, y_grid)
    te_report = class_re(Y_test, y_grid, target_names=target_names, output_dict=True)

    all_scores = score_tuple(*[test_confusion, tr_report, te_report, train_matthews,
                               test_matthews, train_confusion])

    return all_scores

def to_dataframe(score_list, name):
    """ A function that transforms the data into dataframes"""
    matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])

    # Taking the confusion matrix
    test_confusion = matrix(*score_list.test_confusion.ravel())
    training_confusion = matrix(*score_list.train_confusion.ravel())

    # Separating confusion matrix into individual elements
    test_true_n = test_confusion.true_n
    test_false_p = test_confusion.false_p
    test_false_n = test_confusion.false_n
    test_true_p = test_confusion.true_p

    training_true_n = training_confusion.true_n
    training_false_p = training_confusion.false_p
    training_false_n = training_confusion.false_n
    training_true_p = training_confusion.true_p

    # coonstructing the dataframe
    dataframe = pd.DataFrame([test_true_n, test_true_p, test_false_p, test_false_n, training_true_n,
                                 training_true_p, training_false_p, training_false_n, score_list.test_mat,
                                 score_list.train_mat])


    dataframe = dataframe.transpose()

    dataframe.columns = ["test_tn", "test_tp", "test_fp", "test_fn", "train_tn", "train_tp",
                           "train_fp", "train_fn", "test_Mat", "train_Mat", ]
    dataframe.index = name

    te_report = pd.DataFrame(score_list.te_report).transpose()
    tr_report = pd.DataFrame(score_list.tr_report).transpose()
    te_report.columns = [f"{x}_{''.join(name)}" for x in te_report.columns]
    tr_report.columns = [f"{x}_{''.join(name)}" for x in tr_report.columns]

    return dataframe, te_report, tr_report

def fit():
    """A function that trains the classifiers and ensembles them"""
    # the features and the labels
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE")
    Y = VHSE["label1"].copy()
    X_knn = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="random_30")

    if X_knn.isnull().values.any():
        X_knn.dropna(axis=1, inplace=True)
        X_knn.drop(["go"], axis=1, inplace=True)

    model_dict = {20: KNN(n_neighbors=7, p=5, metric="minkowski", n_jobs=-1), 40: KNN(n_neighbors=5, p=5, metric="minkowski", n_jobs=-1),
                  70: KNN(n_neighbors=7, p=1, metric="minkowski", n_jobs=-1), 80: KNN(n_neighbors=4, p=4, metric="minkowski", n_jobs=-1),
                  90: KNN(n_neighbors=9, p=1, metric="minkowski", n_jobs=-1)}

    for states in [20, 40, 70, 80, 90]:
        # split and train
        transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn, Y, states)
        model_dict[states].fit(transformed_x_knn, Y_train_knn)

    return model_dict


def predict(fitted, test_x, transformed_x):
    """Using fitted models it makes predictions"""
    preditions_train = {}
    predictions_test = {}
    for states in [20, 40, 70, 80, 90]:
        # predict on X_test
        predictions_test[states] = fitted[states].predict(test_x)
        # predict on X_train
        preditions_train[states] = fitted[states].predict(transformed_x)

    return predictions_test, preditions_train


def get_scores(Y_test_svc, Y_train_svc, test_pred, train_pred):
    """Converts the scores into dataframes"""
    # ensembles the predictions
    vote_2_test, index2_test = vote(test_pred[20], test_pred[40], test_pred[70])
    vote_2_train, index2_train = vote(train_pred[20], train_pred[40], train_pred[70])
    vote_1_test, index1_test = vote(test_pred[20], test_pred[40])
    vote_1_train, index1_train = vote(train_pred[20], train_pred[40])
    vote_3_test, index3_test = vote(test_pred[20], test_pred[40], test_pred[70], test_pred[80])
    vote_3_train, index3_train = vote(train_pred[20], train_pred[40], train_pred[70], train_pred[80])
    vote_4_test, index4_test = vote(test_pred[20], test_pred[40], test_pred[70], test_pred[80], test_pred[90])
    vote_4_train, index4_train = vote(train_pred[20], train_pred[40], train_pred[70], train_pred[80], train_pred[90])

    # generating the scores
    individual_scores = []
    for states in [20, 40, 70, 80, 90]:
        score = print_score(Y_test_svc, test_pred[states], train_pred[states], Y_train_svc)
        individual_scores.append(score)

    ensemble2_purged = print_score(Y_test_svc, vote_2_test, vote_2_train, Y_train_svc, index2_test, index2_train,
                                   mode=1)
    ensemble1_purged = print_score(Y_test_svc, vote_1_test, vote_1_train, Y_train_svc, index1_test, index1_train,
                                   mode=1)
    ensemble3_purged = print_score(Y_test_svc, vote_3_test, vote_3_train, Y_train_svc, index3_test, index3_train,
                                   mode=1)
    ensemble4_purged = print_score(Y_test_svc, vote_4_test, vote_4_train, Y_train_svc, index4_test, index4_train,
                                   mode=1)

    # put all the sores into dataframe
    dataframe_20, te_report_20, tr_report_20 = to_dataframe(individual_scores[0], ["knn_20"])
    dataframe_40, te_report_40, tr_report_40 = to_dataframe(individual_scores[1], ["knn_40"])
    dataframe_70, te_report_70, tr_report_70 = to_dataframe(individual_scores[2], ["knn_70"])
    dataframe_80, te_report_80, tr_report_80 = to_dataframe(individual_scores[3], ["knn_80"])
    dataframe_90, te_report_90, tr_report_90 = to_dataframe(individual_scores[4], ["knn_90"])
    dataframe_ense1, te_report_ense1, tr_report_ense1 = to_dataframe(ensemble1_purged, ["ensemble1"])
    dataframe_ense2, te_report_ense2, tr_report_ense2 = to_dataframe(ensemble2_purged, ["ensemble2"])
    dataframe_ense3, te_report_ense3, tr_report_ense3 = to_dataframe(ensemble3_purged, ["ensemble3"])
    dataframe_ense4, te_report_ense4, tr_report_ense4 = to_dataframe(ensemble4_purged, ["ensemble4"])

    # join the dataframes
    all_data = pd.concat([dataframe_20, dataframe_40, dataframe_70, dataframe_80, dataframe_90, dataframe_ense1,
                          dataframe_ense2, dataframe_ense3, dataframe_ense4], axis=0)
    all_te_report = pd.concat([te_report_20, te_report_40, te_report_70, te_report_80, te_report_90,
                               te_report_ense1, te_report_ense2, te_report_ense3, te_report_ense4], axis=1)
    all_tr_report = pd.concat([tr_report_20, tr_report_40, tr_report_70, tr_report_80, tr_report_90,
                               tr_report_ense1, tr_report_ense2, tr_report_ense3, tr_report_ense4], axis=1)

    return all_data, all_te_report, all_tr_report

def writing(dataset1, te_report, tr_report, sheet_name, row=0):
    """Writes to excel"""
    if not path.exists(f"ensemble_scores/intramodel_knn.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"ensemble_scores/intramodel_knn.xlsx")

    book = load_workbook('ensemble_scores/intramodel_knn.xlsx')
    writer = pd.ExcelWriter('ensemble_scores/intramodel_knn.xlsx', engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    tr_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3, )
    te_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3+len(tr_report.index)+3)

    writer.save()
    writer.close()

def run_esterase(name):
    """run all the script"""
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE")
    Y = VHSE["label1"].copy()
    X_knn = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="random_30")

    if X_knn.isnull().values.any():
        X_knn.dropna(axis=1, inplace=True)
        X_knn.drop(["go"], axis=1, inplace=True)


    # generating the predictions
    fitted_models = fit()
    for states in [20, 40, 70, 80, 90]:
        transformed_x_knn, test_x_knn, Y_train_knn, Y_test_knn, X_train_knn, X_test_knn = split_transform(X_knn, Y, states)
        test_pred, train_pred = predict(fitted_models, test_x_knn, transformed_x_knn)
    # get the scores
        all_data, all_te_report, all_tr_report = get_scores(Y_test_knn, Y_train_knn, test_pred, train_pred)

        writing(all_data, all_te_report, all_tr_report, f"{name}_{states}")

run_esterase("knn")