from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier as KNN
from sklearn.metrics import matthews_corrcoef, confusion_matrix
from sklearn.metrics import classification_report as class_re
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.linear_model import RidgeClassifier as RIDGE
from sklearn.svm import SVC
import pandas as pd
from collections import namedtuple
import numpy as np
from os import path

def split_transform(X, Y, states=20):
    """Given X and Y returns a split and scaled version of them"""
    scaling = MinMaxScaler()
    esterase = ['EH51(22)', 'EH75(16)', 'EH46(23)', 'EH98(11)', 'EH49(23)']

    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=states, stratify=Y)

    X_train = X_train.loc[[x for x in X_train.index if x not in esterase]]
    X_test = X_test.loc[[x for x in X_test.index if x not in esterase]]
    Y_train = Y_train.loc[[x for x in Y_train.index if x not in esterase]]
    Y_test = Y_test.loc[[x for x in Y_test.index if x not in esterase]]

    transformed_x = scaling.fit_transform(X_train)
    transformed_x = pd.DataFrame(transformed_x)
    transformed_x.index = X_train.index
    transformed_x.columns = X_train.columns

    test_x = scaling.transform(X_test)
    test_x = pd.DataFrame(test_x)
    test_x.index = X_test.index
    test_x.columns = X_test.columns

    return transformed_x, test_x, Y_train, Y_test, X_train, X_test

def vote(pred1, pred2, pred3=None):
    """Hard voting for the ensembles"""
    vote_ = []
    index = []
    if pred3 is None:
        mean = np.mean([pred1, pred2], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            else:
                vote_.append(pred2[s])
                index.append(s)
    else:
        mean = np.mean([pred1, pred2, pred3], axis=0)
        for s, x in enumerate(mean):
            if x == 1 or x == 0:
                vote_.append(int(x))
            elif x > 0.5:
                vote_.append(1)
                index.append(s)
            else:
                vote_.append(0)
                index.append(s)

    return mean, vote_, index

def print_score(Y_test, y_grid, train_predicted, Y_train, test_index=None, train_index=None, mode=None):
    """ The function prints the scores of the models and the prediction performance """
    score_tuple = namedtuple("scores", ["test_confusion", "tr_report", "te_report",
                                        "train_mat", "test_mat", "train_confusion"])

    target_names = ["class 0", "class 1"]

    # looking at the scores of those predicted by al 3 of them
    if mode:
        Y_test = Y_test.iloc[[x for x in range(len(Y_test)) if x not in test_index]]
        Y_train = Y_train.iloc[[x for x in range(len(Y_train)) if x not in train_index]]
        y_grid = [y_grid[x] for x in range(len(y_grid)) if x not in test_index]
        train_predicted = [train_predicted[x] for x in range(len(train_predicted)) if x not in train_index]

    # Training scores
    train_confusion = confusion_matrix(Y_train, train_predicted)
    train_matthews = matthews_corrcoef(Y_train, train_predicted)
    # print(f"Y_train : {Y_train}, predicted: {train_predicted}")
    tr_report = class_re(Y_train, train_predicted, target_names=target_names, output_dict=True)

    # Test metrics
    test_confusion = confusion_matrix(Y_test, y_grid)
    test_matthews = matthews_corrcoef(Y_test, y_grid)
    te_report = class_re(Y_test, y_grid, target_names=target_names, output_dict=True)

    all_scores = score_tuple(*[test_confusion, tr_report, te_report, train_matthews,
                               test_matthews, train_confusion])

    return all_scores

def to_dataframe(score_list, name):
    """ A function that transforms the data into dataframes"""
    matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])

    # Taking the confusion matrix
    test_confusion = matrix(*score_list.test_confusion.ravel())
    training_confusion = matrix(*score_list.train_confusion.ravel())

    # Separating confusion matrix into individual elements
    test_true_n = test_confusion.true_n
    test_false_p = test_confusion.false_p
    test_false_n = test_confusion.false_n
    test_true_p = test_confusion.true_p

    training_true_n = training_confusion.true_n
    training_false_p = training_confusion.false_p
    training_false_n = training_confusion.false_n
    training_true_p = training_confusion.true_p

    # coonstructing the dataframe
    dataframe = pd.DataFrame([test_true_n, test_true_p, test_false_p, test_false_n, training_true_n,
                                 training_true_p, training_false_p, training_false_n, score_list.test_mat,
                                 score_list.train_mat])


    dataframe = dataframe.transpose()

    dataframe.columns = ["test_tn", "test_tp", "test_fp", "test_fn", "train_tn", "train_tp",
                           "train_fp", "train_fn", "test_Mat", "train_Mat", ]
    dataframe.index = name

    te_report = pd.DataFrame(score_list.te_report).transpose()
    tr_report = pd.DataFrame(score_list.tr_report).transpose()
    te_report.columns = [f"{x}_{''.join(name)}" for x in te_report.columns]
    tr_report.columns = [f"{x}_{''.join(name)}" for x in tr_report.columns]

    return dataframe, te_report, tr_report

def fit():
    """A function that trains the classifiers and ensembles them"""
    # the features and the labels
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE")
    Y = VHSE["label1"].copy()
    X_svc = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="ch2_20")

    if X_svc.isnull().values.any():
        X_svc.dropna(axis=1, inplace=True)
        X_svc.drop(["go"], axis=1, inplace=True)

    # named_tuples
    models = namedtuple("models", ["svc", "ridge", "knn"])
    test = namedtuple("test_samples", ["x_svc", "y_svc",  "x_test_svc"])
    train = namedtuple("train_samples", ["svc_x", "svc_y", "x_train_svc"])

    # split and train
    transformed_x_svc, test_x_svc, Y_train_svc, Y_test_svc, X_train_svc, X_test_svc = split_transform(X_svc, Y)

    # the 3 algorithms
    svc = SVC(C=0.31, kernel="rbf", gamma=0.91)
    knn = KNN(n_neighbors=10, p=5, metric="minkowski", n_jobs=-1)
    ridge = RIDGE(alpha=8, random_state=0)

    # fit the 3 algorithms
    svc.fit(transformed_x_svc, Y_train_svc)
    ridge.fit(transformed_x_svc, Y_train_svc)
    knn.fit(transformed_x_svc, Y_train_svc)

    # save in namedtuples
    fitted_models = models(*[svc, ridge, knn])
    test_sample = test(*[test_x_svc, Y_test_svc, X_test_svc])
    train_sample = train(*[transformed_x_svc, Y_train_svc, X_train_svc])

    return fitted_models, test_sample, train_sample

def predict(fitted, test_x_svc, transformed_x_svc):
    """Using fitted models it makes predictions"""
    # name_tuples
    predictions = namedtuple("predictions", ["svc", "ridge", "knn"])

    # predict on X_test
    y_svc_pred = fitted.svc.predict(test_x_svc)
    y_ridge_pred = fitted.ridge.predict(test_x_svc)
    y_knn_pred = fitted.knn.predict(test_x_svc)

    # predict on X_train
    train_y_svc = fitted.svc.predict(transformed_x_svc)
    train_y_ridge = fitted.ridge.predict(transformed_x_svc)
    train_y_knn = fitted.knn.predict(transformed_x_svc)

    test_pred = predictions(*[y_svc_pred, y_ridge_pred, y_knn_pred])
    train_pred = predictions(*[train_y_svc, train_y_ridge, train_y_knn])

    return test_pred, train_pred

def get_scores(Y_test_svc, Y_train_svc, test_pred, train_pred):
    """Converts the scores into dataframes"""
    # ensembles the predictions
    mean1_test, vote_1_test, index1_test = vote(test_pred.svc, test_pred.ridge, test_pred.knn)
    mean1_train, vote_1_train, index1_train = vote(train_pred.svc, train_pred.ridge, train_pred.knn)
    mean2_test, vote_2_test, index2_test = vote(test_pred.svc, test_pred.ridge)
    mean2_train, vote_2_train, index2_train = vote(train_pred.svc, train_pred.ridge)

    # generating the scores
    scores_svc = print_score(Y_test_svc, test_pred.svc, train_pred.svc, Y_train_svc)
    scores_ridge = print_score(Y_test_svc, test_pred.ridge, train_pred.ridge, Y_train_svc)
    knn_scores = print_score(Y_test_svc, test_pred.knn, train_pred.knn, Y_train_svc)
    ensemble2_score = print_score(Y_test_svc, vote_2_test, vote_2_train, Y_train_svc)
    ensemble2_purged = print_score(Y_test_svc, vote_2_test, vote_2_train, Y_train_svc, index2_test, index2_train, mode=1)
    ensemble1_score = print_score(Y_test_svc, vote_1_test, vote_1_train, Y_train_svc)
    ensemble1_purged = print_score(Y_test_svc, vote_1_test, vote_1_train, Y_train_svc, index1_test, index1_train, mode=1)

    # put all the sores into dataframe
    dataframe_svc, te_report_svc, tr_report_svc = to_dataframe(scores_svc, ["svc"])
    dataframe_ridge, te_report_ridge, tr_report_ridge = to_dataframe(scores_ridge, ["ridge"])
    dataframe_knn, te_report_knn, tr_report_knn = to_dataframe(knn_scores, ["knn"])
    dataframe_ense1, te_report_ense1, tr_report_ense1 = to_dataframe(ensemble1_score, ["ensemble1"])
    dataframe_ense2, te_report_ense2, tr_report_ense2 = to_dataframe(ensemble2_score, ["ensemble2"])
    dataframe_ense1_purged, te_report_ense1_purged, tr_report_ense1_purged = to_dataframe(ensemble1_purged, ["ensemble1_purged"])
    dataframe_ense2_purged, te_report_ense2_purged, tr_report_ense2_purged = to_dataframe(ensemble2_purged, ["ensemble2_purged"])

    # join the dataframes
    all_data = pd.concat([dataframe_svc, dataframe_ridge, dataframe_knn, dataframe_ense1, dataframe_ense2,
                          dataframe_ense1_purged, dataframe_ense2_purged], axis=0)
    all_te_report = pd.concat([te_report_svc, te_report_ridge, te_report_knn, te_report_ense1, te_report_ense2,
                               te_report_ense1_purged, te_report_ense2_purged], axis=1)
    all_tr_report = pd.concat([tr_report_svc, tr_report_ridge, tr_report_knn, tr_report_ense1, tr_report_ense2,
                               tr_report_ense1_purged, tr_report_ense2_purged], axis=1)

    return all_data, all_te_report, all_tr_report

def writing(dataset1, te_report, tr_report, sheet_name, row=0):
    """Writes to excel"""
    if not path.exists(f"ensemble_scores/ensemble_ch2.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=f"ensemble_scores/ensemble_ch2.xlsx")

    book = load_workbook('ensemble_scores/ensemble_ch2.xlsx')
    writer = pd.ExcelWriter('ensemble_scores/ensemble_ch2.xlsx', engine='openpyxl')

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    dataset1.to_excel(writer, sheet_name=f'{sheet_name}', startrow=row)
    tr_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3, )
    te_report.to_excel(writer, sheet_name=f'{sheet_name}', startrow=len(dataset1.index)+3+len(tr_report.index)+3)

    writer.save()
    writer.close()

def run_esterase(name):
    """run all the script"""
    VHSE = pd.read_excel("sequences.xlsx", index_col=0, sheet_name="VHSE")
    Y = VHSE["label1"].copy()
    X_svc = pd.read_excel("esterase_binary.xlsx", index_col=0, sheet_name="ch2_20")

    if X_svc.isnull().values.any():
        X_svc.dropna(axis=1, inplace=True)
        X_svc.drop(["go"], axis=1, inplace=True)

    # generating the predictions
    fitted_models, test_sample, train_sample = fit()
    for states in [20, 40, 70, 80, 90]:
        transformed_x_svc, test_x_svc, Y_train_svc, Y_test_svc, X_train_svc, X_test_svc = split_transform(X_svc, Y, states)
        test_pred, train_pred = predict(fitted_models, test_x_svc, transformed_x_svc)
    # get the scores
        all_data, all_te_report, all_tr_report = get_scores(Y_test_svc, Y_train_svc, test_pred, train_pred)

        writing(all_data, all_te_report, all_tr_report, f"{name}_{states}")

"""Fosfatases"""
def read_fosfa():
    """read the features from fosfatases"""
    data = pd.read_excel("../../phosphatase/fosfatasa_cleaned.xlsx", index_col=1)
    ID = data.index

    # features from ifeatures
    amphy = pd.read_csv("../../phosphatase/ifeature/download/APAAC.tsv", sep="\t", index_col=0)
    amphy.index = ID
    pse_AAC = pd.read_csv("../../phosphatase/ifeature/download/PAAC.tsv", sep="\t", index_col=0)
    pse_AAC.index = ID
    comp_space_aa_group_pairs = pd.read_csv("../../phosphatase/ifeature/download/CKSAAGP.tsv", sep="\t", index_col=0)
    comp_space_aa_group_pairs.index = ID
    distribution = pd.read_csv("../../phosphatase/ifeature/download/CTDD.tsv", sep="\t", index_col=0)
    distribution.index = ID
    composition = pd.read_csv("../../phosphatase/ifeature/download/CTDC.tsv", sep="\t", index_col=0)
    composition.index = ID
    transition = pd.read_csv("../../phosphatase/ifeature/download/CTDT.tsv", sep="\t", index_col=0)
    transition.index = ID
    descriptors = pd.read_csv("../../phosphatase/ifeature/download/descriptors.tsv", sep="\t", index_col=0)
    descriptors.columns = [f"{y}_{x}" for x, y in enumerate(descriptors.columns)]
    descriptors.index = ID
    group_di_comp = pd.read_csv("../../phosphatase/ifeature/download/GDPC.tsv", sep="\t", index_col=0)
    group_di_comp.index = ID
    geary = pd.read_csv("../../phosphatase/ifeature/download/Geary.tsv", sep="\t", index_col=0)
    geary.columns = [f"{x}_geary" for x in geary.columns]
    geary.index = ID
    group_tri_comp = pd.read_csv("../../phosphatase/ifeature/download/GTPC.tsv", sep="\t", index_col=0)
    group_tri_comp.index = ID
    moran = pd.read_csv("../../phosphatase/ifeature/download/Moran.tsv", sep="\t", index_col=0)
    moran.columns = [f"{x}_moran" for x in moran.columns]
    moran.index = ID
    broto = pd.read_csv("../../phosphatase/ifeature/download/NMBroto.tsv", sep="\t", index_col=0)
    broto.columns = [f"{x}_broto" for x in broto.columns]
    broto.index = ID
    quasi_order = pd.read_csv("../../phosphatase/ifeature/download/QSOrder.tsv", sep="\t", index_col=0)
    quasi_order.index = ID
    sequence_order = pd.read_csv("../../phosphatase/ifeature/download/SOCNumber.tsv", sep="\t", index_col=0)
    sequence_order.index = ID
    group_aa_comp = pd.read_csv("../../phosphatase/ifeature/download/GAAC.tsv", sep="\t", index_col=0)
    group_aa_comp.index = ID
    k_space_ct = pd.read_csv("../../phosphatase/ifeature/download/KSCtriad.tsv", sep="\t", index_col=0)
    k_space_ct.index = ID
    ct = pd.read_csv("../../phosphatase/ifeature/download/Ctriad.tsv", sep="\t", index_col=0)
    ct.index = ID

    # features from possum
    ab_pssm = pd.read_csv("../../phosphatase/possum/default/ab_pssm.csv")
    ab_pssm.index = ID
    d_fpssm = pd.read_csv("../../phosphatase/possum/default/d_fpssm.csv")
    d_fpssm.index = ID
    dp_pssm = pd.read_csv("../../phosphatase/possum/default/dp_pssm.csv")
    dp_pssm.index = ID
    dpc_pssm = pd.read_csv("../../phosphatase/possum/default/dpc_pssm.csv")
    dpc_pssm.index = ID
    edp = pd.read_csv("../../phosphatase/possum/default/edp.csv")
    edp.index = ID
    pssm_ac = pd.read_csv("../../phosphatase/possum/default/pssm_ac.csv")
    pssm_ac.index = ID
    pssm_cc = pd.read_csv("../../phosphatase/possum/default/pssm_cc.csv")
    pssm_cc.index = ID
    pssm_composition = pd.read_csv("../../phosphatase/possum/default/pssm_composition.csv")
    pssm_composition.index = ID
    rpm_pssm = pd.read_csv("../../phosphatase/possum/default/rpm_pssm.csv")
    rpm_pssm.index = ID
    rpssm = pd.read_csv("../../phosphatase/possum/default/rpssm.csv")
    rpssm.index = ID
    s_fpssm = pd.read_csv("../../phosphatase/possum/default/s_fpssm.csv")
    s_fpssm.index = ID
    tpc = pd.read_csv("../../phosphatase/possum/default/tpc.csv")
    tpc.index = ID
    tri_gram_pssm = pd.read_csv("../../phosphatase/possum/default/tri_gram_pssm.csv")
    tri_gram_pssm.index = ID
    eedp = pd.read_csv("../../phosphatase/possum/default/eedp.csv")
    eedp.index = ID
    bigrams_pssm = pd.read_csv("../../phosphatase/possum/default/k_separated_bigrams_pssm.csv")
    bigrams_pssm.index = ID

    pse_pssm_1 = pd.read_csv("../../phosphatase/possum/default/pse_pssm.csv")
    index = pse_pssm_1.columns
    index_1 = [f"{x}_1" for x in index]
    index_2 = [f"{x}_2" for x in index]
    index_3 = [f"{x}_3" for x in index]
    pse_pssm_1.index = ID
    pse_pssm_1.columns = index_1

    pse_pssm_2 = pd.read_csv("../../phosphatase/possum/pse_pssm2/pse_pssm.csv")
    pse_pssm_2.index = ID
    pse_pssm_2.columns = index_2

    pse_pssm_3 = pd.read_csv("../../phosphatase/possum/pse_pssm3/pse_pssm.csv")
    pse_pssm_3.index = ID
    pse_pssm_3.columns = index_3

    smoothed_pssm_7 = pd.read_csv("../../phosphatase/possum/default/smoothed_pssm.csv")

    index_smoo = smoothed_pssm_7.columns
    index_smoo_5 = [f"{x}_5" for x in index_smoo]
    index_smoo_7 = [f"{x}_7" for x in index_smoo]
    index_smoo_9 = [f"{x}_9" for x in index_smoo]

    smoothed_pssm_7.index = ID
    smoothed_pssm_7.columns = index_smoo_7

    smoothed_pssm_5 = pd.read_csv("../../phosphatase/possum/pse_pssm2/smoothed_pssm.csv")
    smoothed_pssm_5.index = ID
    smoothed_pssm_5.columns = index_smoo_5

    smoothed_pssm_9 = pd.read_csv("../../phosphatase/possum/pse_pssm3/smoothed_pssm.csv")
    smoothed_pssm_9.index = ID
    smoothed_pssm_9.columns = index_smoo_9

    aac_pssm = pd.read_csv("../../phosphatase/possum/default/aac_pssm.csv")
    aac_pssm.index = ID
    aac_pssm.head()

    # concatenate features from ifeatures
    all_data = pd.concat([amphy, transition, group_di_comp, geary, group_tri_comp,
                          moran, pse_AAC, comp_space_aa_group_pairs, distribution, descriptors,
                          composition, broto, group_aa_comp, sequence_order, quasi_order, ct, k_space_ct], axis=1)

    # concatenate features from possum
    feature = [aac_pssm, ab_pssm, d_fpssm, dp_pssm, dpc_pssm, edp, eedp, bigrams_pssm, pssm_ac, pssm_cc,
               pssm_composition, rpm_pssm, rpssm, s_fpssm, tpc, tri_gram_pssm, pse_pssm_1, pse_pssm_2,
               pse_pssm_3, smoothed_pssm_5, smoothed_pssm_7, smoothed_pssm_9]
    everything = pd.concat(feature, axis=1)

    # sort all the features
    all = pd.concat([all_data, everything], axis=1)
    all["hits"] = data["Hits"]
    all = all.sort_values(by="hits", axis=0, ascending=False)
    all.drop(["hits"], axis=1, inplace=True)

    return all

def differential_split(sub_ig, label="label"):
    """A function that splits the dataset according to cap type"""
    Y = namedtuple("Y", ["c0_y", "c1_y", "c2_y", "cy_other"])
    X = namedtuple("X", ["c0_x", "c1_x", "c2_x", "cx_other"])

    # reading the labels and the features
    y = pd.read_excel("../../phosphatase/fosfatasa_cleaned.xlsx", index_col=0, sheet_name="sorted")

    sub_ig_new_concat = pd.concat([sub_ig, y["Cap Type"], y[label]], axis=1)
    # creating X and Y of teh c¡different cap types
    c0_x = sub_ig_new_concat[sub_ig_new_concat["Cap Type"] == "C0"].copy()
    c1_x = sub_ig_new_concat[sub_ig_new_concat["Cap Type"] == "C1"].copy()
    c2_x = sub_ig_new_concat[sub_ig_new_concat["Cap Type"] == "C2"].copy()
    cx_other = sub_ig_new_concat[sub_ig_new_concat["Cap Type"] == "C1+C2"].copy()

    c0_y = c0_x[label].copy()
    c1_y = c1_x[label].copy()
    c2_y = c2_x[label].copy()
    cy_other = cx_other[label].copy()

    c0_x.drop(["Cap Type", label], inplace=True, axis=1)
    c1_x.drop(["Cap Type", label], inplace=True, axis=1)
    c2_x.drop(["Cap Type", label], inplace=True, axis=1)
    cx_other.drop(["Cap Type", label], inplace=True, axis=1)
    # saving the datasets
    y_list = Y(*[c0_y, c1_y, c2_y, cy_other])
    x_list = X(*[c0_x, c1_x, c2_x, cx_other])

    return x_list, y_list

def fosfatase_split(x_list, y_list, state=20):
    """given the dataset it scales it """
    X_train_c0, X_test_c0, Y_train_c0, Y_test_c0 = train_test_split(x_list.c0_x, y_list.c0_y, test_size=0.2,
                                                                    random_state=state)
    X_train_c1, X_test_c1, Y_train_c1, Y_test_c1 = train_test_split(x_list.c1_x, y_list.c1_y, test_size=0.2,
                                                                    random_state=state, stratify=y_list.c1_y)
    X_train_c2, X_test_c2, Y_train_c2, Y_test_c2 = train_test_split(x_list.c2_x, y_list.c2_y, test_size=0.2,
                                                                    random_state=state, stratify=y_list.c2_y)

    X_train = pd.concat([X_train_c0, X_train_c1, X_train_c2, x_list.cx_other], axis=0)
    X_test = pd.concat([X_test_c0, X_test_c1, X_test_c2], axis=0)
    Y_train = pd.concat([Y_train_c0, Y_train_c1, Y_train_c2, y_list.cy_other], axis=0)
    Y_test = pd.concat([Y_test_c0, Y_test_c1, Y_test_c2], axis=0)

    fosfatase = [501030, 508415, 508252, 900101, 900184, 508325, 501257]
    X_train = X_train.loc[[x for x in X_train.index if x not in fosfatase]]
    X_test = X_test.loc[[x for x in X_test.index if x not in fosfatase]]
    Y_train = Y_train.loc[[x for x in Y_train.index if x not in fosfatase]]
    Y_test = Y_test.loc[[x for x in Y_test.index if x not in fosfatase]]

    return X_train, X_test, Y_train, Y_test

def transforming(X_train, X_test):
    """transforms the features from fosfatases"""
    scaling = MinMaxScaler()
    transformed_x = scaling.fit_transform(X_train)
    transformed_x = pd.DataFrame(transformed_x)
    transformed_x.index = X_train.index
    transformed_x.columns = X_train.columns

    test_x = scaling.transform(X_test)
    test_x = pd.DataFrame(test_x)
    test_x.index = X_test.index
    test_x.columns = X_test.columns

    return transformed_x, test_x

def predict_fosfatase(name):
    """Used trained models to make predictions on phosphatases"""
    # generate fosfatase features
    all = read_fosfa()
    x_list, y_list = differential_split(all)

    # keep the features that are the same as in knn
    fitted_models, test_sample, train_sample = fit()

    for states in [40, 70, 80, 90]:
        X_train, X_test, Y_train, Y_test = fosfatase_split(x_list, y_list, states)

        train_x_svc = X_train[train_sample.x_train_svc.columns]
        train_x_knn = X_train[train_sample.x_train_knn.columns]
        test_x_svc = X_test[test_sample.x_test_svc.columns]
        test_x_knn = X_test[test_sample.x_test_knn.columns]

        transformed_x_svc, test_x_svc = transforming(train_x_svc, test_x_svc)
        transformed_x_knn, test_x_knn = transforming(train_x_knn, test_x_knn)

        test_pred, train_pred = predict(fitted_models, test_x_svc, test_x_knn, transformed_x_svc,
                                    transformed_x_knn)
        # get the scores
        all_data, all_te_report, all_tr_report = get_scores(Y_test, Y_test, Y_train,
                                                        Y_train, test_pred, train_pred)

        writing(all_data, all_te_report, all_tr_report, f"{name}_{states}")

"""Running the functions"""

run_esterase("ch2")